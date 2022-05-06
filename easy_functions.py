#!/usr/bin/env python3

#======================================================
# Source Modules
#======================================================
from collections import OrderedDict
from openpyxl import load_workbook
from ordered_set import OrderedSet
from textwrap import fill
import ast
import jinja2
import json
import os
import pkg_resources
import platform
import re
import requests
import subprocess
import sys
import stdiomask
import time
import validating

#======================================================
# Log Level - 0 = None, 1 = Class only, 2 = Line
#======================================================
log_level = 2

#======================================================
# Exception Classes
#======================================================
class InsufficientArgs(Exception):
    pass

class ErrException(Exception):
    pass

class InvalidArg(Exception):
    pass

class LoginFailed(Exception):
    pass

#======================================================
# Function to run 'terraform plan' and
# 'terraform apply' in the each folder of the
# Destination Directory.
#======================================================
def apply_aci_terraform(folders):

    print(f'\n-----------------------------------------------------------------------------\n')
    print(f'  Found the Followng Folders with uncommitted changes:\n')
    for folder in folders:
        print(f'  - {folder}')
    print(f'\n  Beginning Terraform Plan and Apply in each folder.')
    print(f'\n-----------------------------------------------------------------------------\n')

    time.sleep(7)

    response_p = ''
    response_a = ''
    for folder in folders:
        path = './%s' % (folder)
        lock_count = 0
        p = subprocess.Popen(['terraform', 'init', '-plugin-dir=../../../terraform-plugins/providers/'],
                             cwd=path,
                             stdout=subprocess.PIPE,
                             stderr=subprocess.STDOUT)
        for line in iter(p.stdout.readline, b''):
            print(line)
            if re.search('does not match configured version', line.decode("utf-8")):
                lock_count =+ 1

        if lock_count > 0:
            p = subprocess.Popen(['terraform', 'init', '-upgrade', '-plugin-dir=../../../terraform-plugins/providers/'], cwd=path)
            p.wait()
        p = subprocess.Popen(['terraform', 'plan', '-out=main.plan'], cwd=path)
        p.wait()
        while True:
            print(f'\n-----------------------------------------------------------------------------\n')
            print(f'  Terraform Plan Complete.  Please Review the Plan and confirm if you want')
            print(f'  to move forward.  "A" to Apply the Plan. "S" to Skip.  "Q" to Quit.')
            print(f'  Current Working Directory: {folder}')
            print(f'\n-----------------------------------------------------------------------------\n')
            response_p = input('  Please Enter ["A", "S" or "Q"]: ')
            if re.search('^(A|S)$', response_p):
                break
            elif response_p == 'Q':
                exit()
            else:
                print(f'\n-----------------------------------------------------------------------------\n')
                print(f'  A Valid Response is either "A", "S" or "Q"...')
                print(f'\n-----------------------------------------------------------------------------\n')

        if response_p == 'A':
            p = subprocess.Popen(['terraform', 'apply', '-parallelism=1', 'main.plan'], cwd=path)
            p.wait()

        while True:
            if response_p == 'A':
                response_p = ''
                print(f'\n-----------------------------------------------------------------------------\n')
                print(f'  Terraform Apply Complete.  Please Review for any errors and confirm if you')
                print(f'  want to move forward.  "M" to Move to the Next Section. "Q" to Quit..')
                print(f'\n-----------------------------------------------------------------------------\n')
                response_a = input('  Please Enter ["M" or "Q"]: ')
            elif response_p == 'S':
                break
            if response_a == 'M':
                break
            elif response_a == 'Q':
                exit()
            else:
                print(f'\n-----------------------------------------------------------------------------\n')
                print(f'  A Valid Response is either "M" or "Q"...')
                print(f'\n-----------------------------------------------------------------------------\n')

#======================================================
# Function to Check the Git Status of the Folders
#======================================================
def check_git_status():
    random_folders = []
    git_path = './'
    result = subprocess.Popen(['python3', '-m', 'git_status_checker'], stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    while(True):
        # returns None while subprocess is running
        retcode = result.poll()
        line = result.stdout.readline()
        line = line.decode('utf-8')
        if re.search(r'M (.*/).*.tf\n', line):
            folder = re.search(r'M (.*/).*.tf\n', line).group(1)
            if not re.search(r'ACI.templates', folder):
                if not folder in random_folders:
                    random_folders.append(folder)
        elif re.search(r'\?\? (.*/).*.tf\n', line):
            folder = re.search(r'\?\? (.*/).*.tf\n', line).group(1)
            if not re.search(r'ACI.templates', folder):
                if not folder in random_folders:
                    random_folders.append(folder)
        elif re.search(r'\?\? (ACI/.*/)\n', line):
            folder = re.search(r'\?\? (ACI/.*/)\n', line).group(1)
            if not (re.search(r'ACI.templates', folder) or re.search(r'\.terraform', folder)):
                if os.path.isdir(folder):
                    folder = [folder]
                    random_folders = random_folders + folder
                else:
                    group_x = [os.path.join(folder, o) for o in os.listdir(folder) if os.path.isdir(os.path.join(folder,o))]
                    random_folders = random_folders + group_x
        if retcode is not None:
            break

    if not random_folders:
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   There were no uncommitted changes in the environment.')
        print(f'   Proceedures Complete!!! Closing Environment and Exiting Script.')
        print(f'\n-----------------------------------------------------------------------------\n')
        exit()

    strict_folders = []
    folder_order = ['Access', 'System', 'Tenant_common', 'Tenant_infra', 'Tenant_mgmt', 'Fabric', 'Admin', 'VLANs', 'Tenant_infra',]
    for folder in folder_order:
        for fx in random_folders:
            if folder in fx:
                if 'ACI' in folder:
                    strict_folders.append(fx)
    for folder in strict_folders:
        if folder in random_folders:
            random_folders.remove(folder)
    for folder in random_folders:
        if 'ACI' in folder:
            strict_folders.append(folder)

    # print(strict_folders)
    return strict_folders

#======================================================
# Function to Count the Number of Keys/Columns
#======================================================
def countKeys(ws, func):
    count = 0
    for i in ws.rows:
        if any(i):
            if str(i[0].value) == func:
                count += 1
    return count

#======================================================
# Function to Create Interface Selectors
#======================================================
def create_selector(ws_sw, ws_sw_row_count, **templateVars):
    print(templateVars['port_count'])
    Port_Selector = ''
    for port in range(1, int(templateVars['port_count']) + 1):
        if port < 10:
            Port_Selector = 'Eth%s-0%s' % (templateVars['module'], port)
        elif port < 100:
            Port_Selector = 'Eth%s-%s' % (templateVars['module'], port)
        elif port > 99:
            Port_Selector = 'Eth%s_%s' % (templateVars['module'], port)
        modp = '%s/%s' % (templateVars['module'],port)
        # Copy the Port Selector to the Worksheet
        data = ['intf_selector',templateVars['Pod_ID'],templateVars['Node_ID'],templateVars['Name'],Port_Selector,modp,'','','','','','']
        ws_sw.append(data)
        rc = '%s:%s' % (ws_sw_row_count, ws_sw_row_count)
        for cell in ws_sw[rc]:
            if ws_sw_row_count % 2 == 0:
                cell.style = 'ws_odd'
            else:
                cell.style = 'ws_even'
        dv1_cell = 'A%s' % (ws_sw_row_count)
        dv2_cell = 'H%s' % (ws_sw_row_count)
        templateVars['dv1'].add(dv1_cell)
        templateVars['dv2'].add(dv2_cell)
        ws_sw_row_count += 1
    return ws_sw_row_count

#======================================================
# Function to Create Static Paths within EPGs
#======================================================
def create_static_paths(wb, wb_sw, row_num, wr_method, dest_dir, dest_file, template, **templateVars):
    wsheets = wb_sw.get_sheet_names()
    tf_file = ''
    for wsheet in wsheets:
        ws = wb_sw[wsheet]
        for row in ws.rows:
            if not (row[12].value == None or row[13].value == None):
                vlan_test = ''
                if re.search('^(individual|port-channel|vpc)$', row[7].value) and (re.search(r'\d+', str(row[12].value)) or re.search(r'\d+', str(row[13].value))):
                    if not row[12].value == None:
                        vlan = row[12].value
                        vlan_test = vlan_range(vlan, **templateVars)
                        if 'true' in vlan_test:
                            templateVars['mode'] = 'native'
                    if not 'true' in vlan_test:
                        templateVars['mode'] = 'regular'
                        if not row[13].value == None:
                            vlans = row[13].value
                            vlan_test = vlan_range(vlans, **templateVars)
                if vlan_test == 'true':
                    templateVars['Pod_ID'] = row[1].value
                    templateVars['Node_ID'] = row[2].value
                    templateVars['Interface_Profile'] = row[3].value
                    templateVars['Interface_Selector'] = row[4].value
                    templateVars['Port'] = row[5].value
                    templateVars['Policy_Group'] = row[6].value
                    templateVars['Port_Type'] = row[7].value
                    templateVars['Bundle_ID'] = row[9].value
                    Site_Group = templateVars['Site_Group']
                    pod = templateVars['Pod_ID']
                    node_id =  templateVars['Node_ID']
                    if templateVars['Port_Type'] == 'vpc':
                        ws_vpc = wb['Inventory']
                        for rx in ws_vpc.rows:
                            if rx[0].value == 'vpc_pair' and int(rx[1].value) == int(Site_Group) and str(rx[4].value) == str(node_id):
                                node1 = templateVars['Node_ID']
                                node2 = rx[5].value
                                templateVars['Policy_Group'] = '%s_vpc%s' % (row[3].value, templateVars['Bundle_ID'])
                                templateVars['tDn'] = 'topology/pod-%s/protpaths-%s-%s/pathep-[%s]' % (pod, node1, node2, templateVars['Policy_Group'])
                                templateVars['Static_Path'] = 'rspathAtt-[topology/pod-%s/protpaths-%s-%s/pathep-[%s]' % (pod, node1, node2, templateVars['Policy_Group'])
                                templateVars['GUI_Static'] = 'Pod-%s/Node-%s-%s/%s' % (pod, node1, node2, templateVars['Policy_Group'])
                                templateVars['Static_descr'] = 'Pod-%s_Nodes-%s-%s_%s' % (pod, node1, node2, templateVars['Policy_Group'])
                                tf_file = './ACI/%s/%s/%s' % (templateVars['Site_Name'], dest_dir, dest_file)
                                read_file = open(tf_file, 'r')
                                read_file.seek(0)
                                static_path_descr = 'resource "aci_epg_to_static_path" "%s_%s_%s"' % (templateVars['App_Profile'], templateVars['EPG'], templateVars['Static_descr'])
                                if not static_path_descr in read_file.read():
                                    create_tf_file(wr_method, dest_dir, dest_file, template, **templateVars)

                    elif templateVars['Port_Type'] == 'port-channel':
                        templateVars['Policy_Group'] = '%s_pc%s' % (row[3].value, templateVars['Bundle_ID'])
                        templateVars['tDn'] = 'topology/pod-%s/paths-%s/pathep-[%s]' % (pod, templateVars['Node_ID'], templateVars['Policy_Group'])
                        templateVars['Static_Path'] = 'rspathAtt-[topology/pod-%s/paths-%s/pathep-[%s]' % (pod, templateVars['Node_ID'], templateVars['Policy_Group'])
                        templateVars['GUI_Static'] = 'Pod-%s/Node-%s/%s' % (pod, templateVars['Node_ID'], templateVars['Policy_Group'])
                        templateVars['Static_descr'] = 'Pod-%s_Node-%s_%s' % (pod, templateVars['Node_ID'], templateVars['Policy_Group'])
                        read_file = open(tf_file, 'r')
                        read_file.seek(0)
                        static_path_descr = 'resource "aci_epg_to_static_path" "%s_%s_%s"' % (templateVars['App_Profile'], templateVars['EPG'], templateVars['Static_descr'])
                        if not static_path_descr in read_file.read():
                            create_tf_file(wr_method, dest_dir, dest_file, template, **templateVars)

                    elif templateVars['Port_Type'] == 'individual':
                        port = 'eth%s' % (templateVars['Port'])
                        templateVars['tDn'] = 'topology/pod-%s/paths-%s/pathep-[%s]' % (pod, templateVars['Node_ID'], port)
                        templateVars['Static_Path'] = 'rspathAtt-[topology/pod-%s/paths-%s/pathep-[%s]' % (pod, templateVars['Node_ID'], port)
                        templateVars['GUI_Static'] = 'Pod-%s/Node-%s/%s' % (pod, templateVars['Node_ID'], port)
                        templateVars['Static_descr'] = 'Pod-%s_Node-%s_%s' % (pod, templateVars['Node_ID'], templateVars['Interface_Selector'])
                        read_file = open(tf_file, 'r')
                        read_file.seek(0)
                        static_path_descr = 'resource "aci_epg_to_static_path" "%s_%s_%s"' % (templateVars['App_Profile'], templateVars['EPG'], templateVars['Static_descr'])
                        if not static_path_descr in read_file.read():
                            create_tf_file(wr_method, dest_dir, dest_file, template, **templateVars)

#======================================================
# Function to Create Terraform auto.tfvars files
#======================================================
def create_tf_file(wr_method, dest_dir, dest_file, template, **templateVars):
    # Make sure the Destination Directory Exists
    dest_dir = './ACI/%s/%s' % (templateVars['Site_Name'], dest_dir)
    if not os.path.isdir(dest_dir):
        mk_dir = 'mkdir -p %s' % (dest_dir)
        os.system(mk_dir)
    # Create File for the Template in the Destination Folder
    tf_file = '%s/%s' % (dest_dir, dest_file)
    wr_file = open(tf_file, wr_method)

    # Render Payload and Write to File
    payload = template.render(templateVars)
    wr_file.write(payload + '\n\n')
    wr_file.close()

#======================================================
# Function to Append the easyDict Dictionary
#======================================================
def easyDict_append(templateVars, **kwargs):
    templateVars = OrderedDict(sorted(templateVars.items()))
    class_type = templateVars['class_type']
    data_type = templateVars['data_type']
    templateVars.pop('data_type')
    if not any(kwargs['site_group'] in d for d in kwargs['easyDict'][class_type][data_type]):
        kwargs['easyDict'][class_type][data_type].append({kwargs['site_group']:[]})
        
    count = 0
    for i in kwargs['easyDict'][class_type][data_type]:
        for k, v in i.items():
            if kwargs['site_group'] == k:
                i[kwargs['site_group']].append(templateVars)
        count += 1
    return kwargs['easyDict']

#======================================================
# Function to Append Subtype easyDict Dictionary
#======================================================
def easyDict_append_subtype(templateVars, **kwargs):
    templateVars = OrderedDict(sorted(templateVars.items()))
    class_type = templateVars['class_type']
    data_type = templateVars['data_type']
    data_subtype = templateVars['data_subtype']
    policy_name = templateVars['policy_name']
    templateVars.pop('class_type')
    templateVars.pop('data_type')
    templateVars.pop('data_subtype')
    templateVars.pop('policy_name')
    count = 0
    templateVars.pop('site_group')
    for item in kwargs['easyDict'][class_type][data_type]:
        for k, v in item.items():
            if k == kwargs['site_group']:
                for i in v:
                    i[data_subtype].append(templateVars)
                    count += 1
    if count == 0 and 'Grp_' in kwargs['site_group']:
        group_id = '%s' % (kwargs['site_group'])
        site_group = ast.literal_eval(os.environ[group_id])
        sites = []
        for x in range(1,16):
            sitex = 'site_%s' % (x)
            if not site_group[sitex] == None:
                sites.append(x)
        count = 0
        for x in sites:
            for item in kwargs['easyDict'][class_type][data_type]:
                for key, value in item.items():
                    if int(key) == int(x):
                        for i in value:
                            if i['name'] == policy_name:
                                i[data_subtype].append(templateVars)
                count += 1

    # Return Dictionary
    return kwargs['easyDict']

#======================================================
# Function to find the Keys for each Worksheet
#======================================================
def findKeys(ws, func_regex):
    func_list = OrderedSet()
    for i in ws.rows:
        if any(i):
            if re.search(func_regex, str(i[0].value)):
                func_list.add(str(i[0].value))
    return func_list

#======================================================
# Function to Create Terraform auto.tfvars files
#======================================================
def findVars(ws, func, rows, count):
    var_list = []
    var_dict = {}
    for i in range(1, rows + 1):
        if (ws.cell(row=i, column=1)).value == func:
            try:
                for x in range(2, 34):
                    if (ws.cell(row=i - 1, column=x)).value:
                        var_list.append(str(ws.cell(row=i - 1, column=x).value))
                    else:
                        x += 1
            except Exception as e:
                e = e
                pass
            break
    vcount = 1
    while vcount <= count:
        var_dict[vcount] = {}
        var_count = 0
        for z in var_list:
            var_dict[vcount][z] = ws.cell(row=i + vcount - 1, column=2 + var_count).value
            var_count += 1
        var_dict[vcount]['row'] = i + vcount - 1
        vcount += 1
    return var_dict

#======================================================
# Function to Merge Easy ACI Repository to Dest Folder
#======================================================
def get_latest_versions(easyDict):
    # Get the Latest Release Tag for the provider-aci repository
    url = f'https://github.com/CiscoDevNet/terraform-provider-aci/tags/'
    r = requests.get(url, stream=True)
    repoVer = 'BLANK'
    stringMatch = False
    while stringMatch == False:
        for line in r.iter_lines():
            toString = line.decode("utf-8")
            if re.search(r'/releases/tag/v(\d+\.\d+\.\d+)\"', toString):
                repoVer = re.search('/releases/tag/v(\d+\.\d+\.\d+)', toString).group(1)
                break
        stringMatch = True
    
    # Set ACI Provider Version
    aci_provider_version = repoVer

    # Get the Latest Release Tag for the provider-mso repository
    url = f'https://github.com/CiscoDevNet/terraform-provider-mso/tags/'
    r = requests.get(url, stream=True)
    repoVer = 'BLANK'
    stringMatch = False
    while stringMatch == False:
        for line in r.iter_lines():
            toString = line.decode("utf-8")
            if re.search(r'/releases/tag/v(\d+\.\d+\.\d+)\"', toString):
                repoVer = re.search('/releases/tag/v(\d+\.\d+\.\d+)', toString).group(1)
                break
        stringMatch = True
    
    # Set NDO Provider Version
    ndo_provider_version = repoVer

    # Get the Latest Release Tag for Terraform
    url = f'https://github.com/hashicorp/terraform/tags'
    r = requests.get(url, stream=True)
    repoVer = 'BLANK'
    stringMatch = False
    while stringMatch == False:
        for line in r.iter_lines():
            toString = line.decode("utf-8")
            if re.search(r'/releases/tag/v(\d+\.\d+\.\d+)\"', toString):
                repoVer = re.search('/releases/tag/v(\d+\.\d+\.\d+)', toString).group(1)
                break
        stringMatch = True

    # Set Terraform Version
    terraform_version = repoVer

     # Get the Latest Release Tag for Nexus Dashboard Orchestrator
    url = f'https://dcappcenter.cisco.com/nexus-dashboard-orchestrator.html'
    r = requests.get(url, stream=True)
    ndoVer = None
    stringMatch = False
    while stringMatch == False:
        for line in r.iter_lines():
            toString = line.decode("utf-8")
            if re.search(r'product-id=\"(\d+)\"', toString):
                ndoVer = re.search(r'product-id=\"(\d+)\"', toString).group(1)
                break
        stringMatch = True

    if ndoVer == None:
        print('\n   Error!!!  Could not find the version of NDO on the dcappcenter.\n')
        exit()
    url = f'https://dcappcenter.cisco.com/rest/V1/product/platforms?id={ndoVer}&approvedOnly=true'
    r = requests.post(url, stream=True)
    ndoVersions = []
    stringMatch = False
    while stringMatch == False:
        for item in r.json():
            ndoVersions.append(item['label'])
        stringMatch = True
    easyDict['latest_versions']['aci_provider_version'] = aci_provider_version
    easyDict['latest_versions']['ndo_provider_version'] = ndo_provider_version
    easyDict['latest_versions']['ndo_versions']['enum'] = ndoVersions
    easyDict['latest_versions']['ndo_versions']['default'] = ndoVersions[0]
    easyDict['latest_versions']['terraform_version'] = terraform_version

    return easyDict

#======================================================
# Function to Get User Password
#======================================================
def get_user_pass():
    print(f'\n-----------------------------------------------------------------------------\n')
    print(f'   Beginning Proceedures to Apply Terraform Resources to the environment')
    print(f'\n-----------------------------------------------------------------------------\n')

    user = input('Enter APIC username: ')
    while True:
        try:
            password = stdiomask.getpass(prompt='Enter APIC password: ')
            break
        except Exception as e:
            print('Something went wrong. Error received: {}'.format(e))

    os.environ['TF_VAR_aciUser'] = '%s' % (user)
    os.environ['TF_VAR_aciPass'] = '%s' % (password)

#======================================================
# Function to Merge Easy ACI Repository to Dest Folder
#======================================================
def merge_easy_aci_repository(easy_jsonData):
    jsonData = easy_jsonData['components']['schemas']['easy_aci']['allOf'][1]['properties']
    
    # Obtain Operating System and Get TF_DEST_DIR variable from Environment
    opSystem = platform.system()
    if os.environ.get('TF_DEST_DIR') is None:
        tfDir = 'Intersight'
    else:
        tfDir = os.environ.get('TF_DEST_DIR')

    # Get All sub-folders from the TF_DEST_DIR
    folders = []
    for root, dirs, files in os.walk(tfDir):
        for name in dirs:
            # print(os.path.join(root, name))
            folders.append(os.path.join(root, name))
    folders.sort()

    # Remove the First Level Folders from the List
    for folder in folders:
        print(f'folder is {folder}')
        if '/' in folder:
            x = folder.split('/')
            if len(x) == 2:
                folders.remove(folder)

    # Get the Latest Release Tag for the terraform-easy-aci repository
    url = f'https://github.com/terraform-cisco-modules/terraform-easy-aci/tags/'
    r = requests.get(url, stream=True)
    repoVer = 'BLANK'
    stringMatch = False
    while stringMatch == False:
        for line in r.iter_lines():
            toString = line.decode("utf-8")
            if re.search('/releases/tag/(\d+\.\d+\.\d+)', toString):
                repoVer = re.search('/releases/tag/(\d+\.\d+\.\d+)', toString).group(1)
                break
        stringMatch = True

    for folder in folders:
        folderVer = "0.0.0"
        # Get the version.txt file if exist to compare to the latest Git Release of the repository
        if opSystem == 'Windows':
            if os.path.isfile(f'{folder}\\version.txt'):
                with open(f'{folder}\\version.txt') as f:
                    folderVer = f.readline().rstrip()
        else:
            if os.path.isfile(f'{folder}/version.txt'):
                with open(f'{folder}/version.txt') as f:
                    folderVer = f.readline().rstrip()
        
        # Determine the Type of Folder. i.e. Is this for Access Policies
        if os.path.isdir(folder):
            if opSystem == 'Windows':
                folder_length = len(folder.split('\\'))
                folder_type = folder.split('\\')[folder_length -1]
            else:
                folder_length = len(folder.split('/'))
                folder_type = folder.split('/')[folder_length -1]
            if re.search('^tenant_', folder_type):
                folder_type = 'tenant'
            
            # Get List of Files to download from jsonData
            files = jsonData['files'][folder_type]
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'\n  Beginning Easy ACI Module Downloads for "{folder}"\n')

            # Run the process to check for the files in the folder and if it doesn't exist download from Github
            for file in files:
                git_url = 'https://raw.github.com/terraform-cisco-modules/terraform-easy-aci/master/modules'
                if opSystem == 'Windows':
                    dest_file = f'{folder}\\{file}'
                else:
                    dest_file = f'{folder}/{file}'
                if not os.path.isfile(dest_file):
                    print(f'  Downloading "{file}"')
                    url = f'{git_url}/{folder_type}/{file}'
                    r = requests.get(url)
                    open(dest_file, 'wb').write(r.content)
                    print(f'  "{file}" Download Complete!\n')
                else:
                    if opSystem == 'Windows':
                        if not os.path.isfile(f'{folder}\\version.txt'):
                            print(f'  Downloading "{file}"')
                            url = f'{git_url}/{folder_type}/{file}'
                            r = requests.get(url)
                            open(dest_file, 'wb').write(r.content)
                            print(f'  "{file}" Download Complete!\n')
                        elif not os.path.isfile(f'{folder}\\version.txt'):
                            print(f'  Downloading "{file}"')
                            url = f'{git_url}/{folder_type}/{file}'
                            r = requests.get(url)
                            open(dest_file, 'wb').write(r.content)
                            print(f'  "{file}" Download Complete!\n')
                        elif os.path.isfile(f'{folder}\\version.txt'):
                            if not folderVer == repoVer:
                                print(f'  Downloading "{file}"')
                                url = f'{git_url}/{folder_type}/{file}'
                                r = requests.get(url)
                                open(dest_file, 'wb').write(r.content)
                                print(f'  "{file}" Download Complete!\n')
                    else:
                        if not os.path.isfile(f'{folder}/version.txt'):
                            print(f'  Downloading "{file}"')
                            url = f'{git_url}/{folder_type}/{file}'
                            r = requests.get(url)
                            open(dest_file, 'wb').write(r.content)
                            print(f'  "{file}" Download Complete!\n')
                        elif not os.path.isfile(f'{folder}/version.txt'):
                            print(f'  Downloading "{file}"')
                            url = f'{git_url}/{folder_type}/{file}'
                            r = requests.get(url)
                            open(dest_file, 'wb').write(r.content)
                            print(f'  "{file}" Download Complete!\n')
                        elif os.path.isfile(f'{folder}/version.txt'):
                            if not folderVer == repoVer:
                                print(f'  Downloading "{file}"')
                                url = f'{git_url}/{folder_type}/{file}'
                                r = requests.get(url)
                                open(dest_file, 'wb').write(r.content)
                                print(f'  "{file}" Download Complete!\n')

            # Create the version.txt file to prevent redundant downloads for the same Github release
            if not os.path.isfile(f'{folder}/version.txt'):
                print(f'* Creating the repo "terraform-easy-aci" version check file\n "{folder}/version.txt"')
                open(f'{folder}/version.txt', 'w').write('%s\n' % (repoVer))
            elif not folderVer == repoVer:
                print(f'* Updating the repo "terraform-easy-aci" version check file\n "{folder}/version.txt"')
                open(f'{folder}/version.txt', 'w').write('%s\n' % (repoVer))

            print(f'\n  Completed Easy IMM Module Downloads for "{folder}"')
            print(f'\n-------------------------------------------------------------------------------------------\n')

    # Loop over the folder list again and create blank auto.tfvars files for anything that doesn't already exist
    for folder in folders:
        if os.path.isdir(folder):
            if opSystem == 'Windows':
                folder_length = len(folder.split('\\'))
                folder_type = folder.split('\\')[folder_length -1]
            else:
                folder_length = len(folder.split('/'))
                folder_type = folder.split('/')[folder_length -1]
            if re.search('^tenant_', folder_type):
                folder_type = 'tenant'
            files = jsonData['files'][folder_type]
            removeList = jsonData['remove_files']
            for xRemove in removeList:
                if xRemove in files:
                    files.remove(xRemove)
            for file in files:
                varFiles = f"{file.split('.')[0]}.auto.tfvars"
                if opSystem == 'Windows':
                    dest_file = f'{folder}\\{varFiles}'
                else:    
                    dest_file = f'{folder}/{varFiles}'
                if not os.path.isfile(dest_file):
                    wr_file = open(dest_file, 'w')
                    x = file.split('.')
                    x2 = x[0].split('_')
                    varList = []
                    for var in x2:
                        var = var.capitalize()
                        varList.append(var)
                    varDescr = ' '.join(varList)
                    varDescr = varDescr + '- Variables'

                    wrString = f'#______________________________________________\n#\n# {varDescr}\n'\
                        '#______________________________________________\n'\
                        '\n%s = {\n}\n' % (file.split('.')[0])
                    wr_file.write(wrString)
                    wr_file.close()

            # Run terraform fmt to cleanup the formating for all of the auto.tfvar files and tf files if needed
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'  Running "terraform fmt" in folder "{folder}",')
            print(f'  to correct variable formatting!')
            print(f'\n-------------------------------------------------------------------------------------------\n')
            p = subprocess.Popen(
                ['terraform', 'fmt', folder],
                stdout = subprocess.PIPE,
                stderr = subprocess.PIPE
            )
            print('Format updated for the following Files:')
            for line in iter(p.stdout.readline, b''):
                line = line.decode("utf-8")
                line = line.strip()
                print(f'- {line}')

#======================================================
# Function to Create Terraform auto.tfvars files
#======================================================
def naming_rule(name_prefix, name_suffix, org):
    if not name_prefix == '':
        name = '%s_%s' % (name_prefix, name_suffix)
    else:
        name = '%s_%s' % (org, name_suffix)
    return name

#======================================================
# Function to Create Terraform auto.tfvars files
#======================================================
def policies_list(policies_list, **templateVars):
    valid = False
    while valid == False:
        print(f'\n-------------------------------------------------------------------------------------------\n')
        if templateVars.get('optional_message'):
            print(templateVars["optional_message"])
        print(f'  {templateVars["policy"]} Options:')
        for i, v in enumerate(policies_list):
            i += 1
            if i < 10:
                print(f'     {i}. {v}')
            else:
                print(f'    {i}. {v}')
        if templateVars["allow_opt_out"] == True:
            print(f'     99. Do not assign a(n) {templateVars["policy"]}.')
        print(f'     100. Create a New {templateVars["policy"]}.')
        print(f'\n-------------------------------------------------------------------------------------------\n')
        policyOption = input(f'Select the Option Number for the {templateVars["policy"]} to Assign to {templateVars["name"]}: ')
        if re.search(r'^[0-9]{1,3}$', policyOption):
            for i, v in enumerate(policies_list):
                i += 1
                if int(policyOption) == i:
                    policy = v
                    valid = True
                    return policy
                elif int(policyOption) == 99:
                    policy = ''
                    valid = True
                    return policy
                elif int(policyOption) == 100:
                    policy = 'create_policy'
                    valid = True
                    return policy

            if int(policyOption) == 99:
                policy = ''
                valid = True
                return policy
            elif int(policyOption) == 100:
                policy = 'create_policy'
                valid = True
                return policy
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'  Error!! Invalid Selection.  Please Select a valid Index from the List.')
            print(f'\n-------------------------------------------------------------------------------------------\n')

#======================================================
# Function to Create Terraform auto.tfvars files
#======================================================
def policies_parse(org, policy_type, policy):
    if os.environ.get('TF_DEST_DIR') is None:
        tfDir = 'Intersight'
    else:
        tfDir = os.environ.get('TF_DEST_DIR')
    policies = []

    opSystem = platform.system()
    if opSystem == 'Windows':
        policy_file = f'.\{tfDir}\{org}\{policy_type}\{policy}.auto.tfvars'
    else:
        policy_file = f'./{tfDir}/{org}/{policy_type}/{policy}.auto.tfvars'
    if os.path.isfile(policy_file):
        if len(policy_file) > 0:
            if opSystem == 'Windows':
                cmd = 'hcl2json.exe %s' % (policy_file)
            else:
                cmd = 'hcl2json %s' % (policy_file)
                # cmd = 'json2hcl -reverse < %s' % (policy_file)
            p = subprocess.run(
                cmd,
                shell=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT
            )
            if 'unable to parse' in p.stdout.decode('utf-8'):
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  !!!! Encountered Error in Attempting to read file !!!!')
                print(f'  - {policy_file}')
                print(f'  Error was:')
                print(f'  - {p.stdout.decode("utf-8")}')
                print(f'\n-------------------------------------------------------------------------------------------\n')
                json_data = {}
                return policies,json_data
            else:
                json_data = json.loads(p.stdout.decode('utf-8'))
                for i in json_data[policy]:
                    policies.append(i)
                return policies,json_data
    else:
        json_data = {}
        return policies,json_data

#======================================================
# Function to validate input for each method
#======================================================
def process_kwargs(required_args, optional_args, **kwargs):
    # Validate all required kwargs passed
    # if all(item in kwargs for item in required_args.keys()) is not True:
    #    error_ = '\n***ERROR***\nREQUIRED Argument Not Found in Input:\n "%s"\nInsufficient required arguments.' % (item)
    #    raise InsufficientArgs(error_)
    row_num = kwargs["row_num"]
    ws = kwargs["ws"]
    error_count = 0
    error_list = []
    for item in required_args:
        if item not in kwargs.keys():
            error_count =+ 1
            error_list += [item]
    if error_count > 0:
        error_ = f'\n\n***Begin ERROR ***\n\nError on Worksheet {ws.title} row {row_num}\n - The Following REQUIRED Key(s) Were Not Found in kwargs: "{error_list}"\n\n****End ERROR****\n'
        raise InsufficientArgs(error_)

    error_count = 0
    error_list = []
    for item in optional_args:
        if item not in kwargs.keys():
            error_count =+ 1
            error_list += [item]
    if error_count > 0:
        error_ = f'\n\n***Begin ERROR***\n\nError on Worksheet {ws.title} row {row_num}\n - The Following Optional Key(s) Were Not Found in kwargs: "{error_list}"\n\n****End ERROR****\n'
        raise InsufficientArgs(error_)

    # Load all required args values from kwargs
    error_count = 0
    error_list = []
    for item in kwargs:
        if item in required_args.keys():
            required_args[item] = kwargs[item]
            if required_args[item] == None:
                error_count =+ 1
                error_list += [item]

    if error_count > 0:
        error_ = f'\n\n***Begin ERROR***\n\nError on Worksheet {ws.title} row {row_num}\n - The Following REQUIRED Key(s) Argument(s) are Blank:\nPlease Validate "{error_list}"\n\n****End ERROR****\n'
        raise InsufficientArgs(error_)

    for item in kwargs:
        if item in optional_args.keys():
            optional_args[item] = kwargs[item]
    # Combine option and required dicts for Jinja template render
    templateVars = {**required_args, **optional_args}
    return(templateVars)

#======================================================
# Function to Add Static Port Bindings to Bridge Domains Terraform Files
#======================================================
def process_workbook(wb, ws, row_num, wr_method, dest_dir, dest_file, template, **templateVars):
    if re.search('Grp_[A-F]', templateVars['site_group']):
        group_id = '%s' % (templateVars['site_group'])
        site_group = ast.literal_eval(os.environ[group_id])
        for x in range(1, 16):
            sitex = 'site_%s' % (x)
            if not site_group[sitex] == None:
                site_id = 'site_id_%s' % (site_group[sitex])
                site_dict = ast.literal_eval(os.environ[site_id])

                # Pull in the Site Workbook
                excel_workbook = '%s_intf_selectors.xlsx' % (templateVars['Site_Name'])
                try:
                    wb_sw = load_workbook(excel_workbook)
                except Exception as e:
                    print(f"Something went wrong while opening the workbook - {excel_workbook}... ABORT!")
                    sys.exit(e)

                # Process the Interface Selectors for Static Port Paths
                create_static_paths(wb, wb_sw, row_num, wr_method, dest_dir, dest_file, template, **templateVars)

    elif re.search(r'\d+', templateVars['Site_Group']):
        site_id = 'site_id_%s' % (templateVars['Site_Group'])
        site_dict = ast.literal_eval(os.environ[site_id])

        # Create templateVars for Site_Name and APIC_URL
        templateVars['Site_Name'] = site_dict.get('Site_Name')
        templateVars['Site_Group'] = site_dict.get('Site_ID')
        templateVars['APIC_URL'] = site_dict.get('APIC_URL')

        # Pull in the Site Workbook
        excel_workbook = '%s_intf_selectors.xlsx' % (templateVars['Site_Name'])
        try:
            wb_sw = load_workbook(excel_workbook)
        except Exception as e:
            print(f"Something went wrong while opening the workbook - {excel_workbook}... ABORT!")
            sys.exit(e)


        # Process the Interface Selectors for Static Port Paths
        create_static_paths(wb, wb_sw, row_num, wr_method, dest_dir, dest_file, template, **templateVars)

    else:
        print(f"\n-----------------------------------------------------------------------------\n")
        print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
        print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
        print(f"\n-----------------------------------------------------------------------------\n")
        exit()

#======================================================
# Function to Determine Port Count on Modules
#======================================================
def query_module_type(row_num, module_type):
    if re.search('^M4', module_type):
        port_count = '4'
    elif re.search('^M6', module_type):
        port_count = '6'
    elif re.search('^M12', module_type):
        port_count = '12'
    elif re.search('X9716D-GX', module_type):
        port_count = '16'
    elif re.search('X9732C-EX', module_type):
        port_count = '32'
    elif re.search('X9736', module_type):
        port_count = '36'
    else:
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   Error on Row {row_num}.  Unknown Module {module_type}')
        print(f'   Please verify Input Information.  Exiting....')
        print(f'\n-----------------------------------------------------------------------------\n')
        exit()
    return port_count

#======================================================
# Function to Determine Port count from Switch Model
#======================================================
# Function to Determine Port count from Switch Model
def query_switch_model(row_num, switch_type):
    modules = ''
    switch_type = str(switch_type)
    if re.search('^9396', switch_type):
        modules = '2'
        port_count = '48'
    elif re.search('^93', switch_type):
        modules = '1'

    if re.search('^9316', switch_type):
        port_count = '16'
    elif re.search('^(93120)', switch_type):
        port_count = '102'
    elif re.search('^(93108|93120|93216|93360)', switch_type):
        port_count = '108'
    elif re.search('^(93180|93240|9348|9396)', switch_type):
        port_count = '54'
    elif re.search('^(93240)', switch_type):
        port_count = '60'
    elif re.search('^9332', switch_type):
        port_count = '34'
    elif re.search('^(9336|93600)', switch_type):
        port_count = '36'
    elif re.search('^9364C-GX', switch_type):
        port_count = '64'
    elif re.search('^9364', switch_type):
        port_count = '66'
    elif re.search('^95', switch_type):
        port_count = '36'
        if switch_type == '9504':
            modules = '4'
        elif switch_type == '9508':
            modules = '8'
        elif switch_type == '9516':
            modules = '16'
        else:
            print(f'\n-----------------------------------------------------------------------------\n')
            print(f'   Error on Row {row_num}.  Unknown Switch Model {switch_type}')
            print(f'   Please verify Input Information.  Exiting....')
            print(f'\n-----------------------------------------------------------------------------\n')
            exit()
    else:
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   Error on Row {row_num}.  Unknown Switch Model {switch_type}')
        print(f'   Please verify Input Information.  Exiting....')
        print(f'\n-----------------------------------------------------------------------------\n')
        exit()
    return modules,port_count

#======================================================
# Function for Processing Loops to auto.tfvars files
#======================================================
def read_easy_jsonData(easy_jsonData, **easyDict):
    jsonData = easy_jsonData['components']['schemas']['easy_aci']['allOf'][1]['properties']
    classes = jsonData['classes']['enum']
    for class_type in classes:
        funcList = jsonData[f'class.{class_type}']['enum']
        for func in funcList:
            sites = []
            site_groups = {}
            for item in easyDict[class_type][func]:
                for k, v in item.items():
                    if re.search('[0-9]+', k):
                        sites.append(k)
                for k, v in item.items():
                    if re.search('Grp_', k):
                        site_groups.update({k:[]})
                        site_group = ast.literal_eval(os.environ[k])
                        gsites = []
                        for kk, vv in site_group.items():
                            if not vv == None and re.search('site_[0-9]+', kk):
                                gsites.append(vv)
                        for site in sites:
                            for gsite in gsites:
                                if int(site) == int(gsite):
                                    site_groups[k].append(site)
            for item in easyDict[class_type][func]:
                loop_count = 1
                for k, v in item.items():
                    dummy_count = 1
                    countlength = len(v)
                    for i in v:
                        # print(i)
                        templateVars = i
                        # print(json.dumps(templateVars, indent=4))
                        # exit()
                        kwargs = {}
                        kwargs['row_num'] = f'{func}_section'
                        kwargs['site_group'] = k
                        kwargs['site_group'] = k
                        kwargs['ws'] = easyDict['wb']['System Settings']
                        
                        # Add Variables for Template Functions
                        templateVars['template_type'] = func
                            
                        if re.search('(apic|bgp_asn)', func):
                            kwargs["template_file"] = 'template_open2.jinja2'
                        else:
                            kwargs["template_file"] = 'template_open.jinja2'
                        if 'bgp' in func:
                            templateVars['policy_type'] = func.replace('_', ' ').upper()
                            kwargs['tfvars_file'] = 'bgp'
                        else:
                            if re.search('aaep_policies', func):
                                kwargs['tfvars_file'] = 'global_policies'
                            elif re.search('(layer3|physical)_domains', func):
                                kwargs['tfvars_file'] = 'domains'
                            elif 'access' in class_type and re.search('policies', func):
                                kwargs['tfvars_file'] = 'interface_policies'
                            elif re.search('leaf_port_group', func):
                                kwargs['tfvars_file'] = 'leaf_interface_policy_groups'
                            elif re.search('spine_port_group', func):
                                kwargs['tfvars_file'] = 'spine_interface_policy_groups'
                            elif 'access' in class_type and re.search('pools', func):
                                kwargs['tfvars_file'] = 'pools'
                            else:
                                kwargs['tfvars_file'] = func
                            x = func.split('_')
                            policyType = ''
                            xcount = 0
                            for i in x:
                                if not i == 'and' and xcount == 0:
                                    policyType = policyType + i.capitalize()
                                elif 'and' in i:
                                    policyType = policyType + ' ' + i
                                else:
                                    policyType = policyType + ' ' + i.capitalize()
                                xcount += 1
                            policyType = policyType.replace('Aaep', 'AAEP')
                            policyType = policyType.replace('Aes', 'AES')
                            policyType = policyType.replace('Apic', 'APIC')
                            policyType = policyType.replace('Cdp', 'CDP')
                            policyType = policyType.replace('Lldp', 'LLDP')
                            policyType = policyType.replace('Radius', 'RADIUS')
                            policyType = policyType.replace('Snmp', 'SNMP')
                            policyType = policyType.replace('Tacacs', 'TACACS+')
                            templateVars['policy_type'] = policyType
                        
                        # Write the Header to the Template File
                        if re.search('aaep|cdp|layer3|group_access|virtual|vlan_p', func) and loop_count == 1:
                            kwargs["initial_write"] = True
                        elif re.search('bgp_rr|policies|port_group|domains|virtual|vlan_p', func):
                            kwargs["initial_write"] = False
                        else:
                            kwargs["initial_write"] = True
                        if re.search('Grp_', k):
                            if not len(site_groups[k]) > 0 and loop_count == 1:
                                write_to_site(templateVars, **kwargs)
                        elif loop_count == 1:
                            write_to_site(templateVars, **kwargs)

                        # Write the template to the Template File
                        kwargs["initial_write"] = False
                        kwargs["template_file"] = f'{func}.jinja2'
                        write_to_site(templateVars, **kwargs)

                        kwargs["initial_write"] = False
                        kwargs["template_file"] = 'template_close.jinja2'

                        
                        if countlength > 1 and countlength != loop_count:
                            dummy_count += 1
                        elif re.search('apic|bgp_asn', func):
                            dummy_count += 1
                        elif re.search('[0-9]+', k):
                            scount = 0
                            for kk, vv in site_groups.items():
                                if k in vv:
                                    scount += 1
                            if scount == 0:
                                write_to_site(templateVars, **kwargs)
                        else:
                            write_to_site(templateVars, **kwargs)
                        loop_count += 1

#======================================================
# Function to Read Excel Workbook Data
#======================================================
def read_in(excel_workbook):
    try:
        wb = load_workbook(excel_workbook)
    except Exception as e:
        print(f"Something went wrong while opening the workbook - {excel_workbook}... ABORT!")
        sys.exit(e)
    return wb

#======================================================
# Function to Add Required Arguments
#======================================================
def required_args_add(args_list, jsonData):
    for i in args_list:
        jsonData['required_args'].update({f'{i}': ''})
        jsonData['optional_args'].pop(i)
    return jsonData

#======================================================
# Function to Add Required Arguments
#======================================================
def required_args_remove(args_list, jsonData):
    for i in args_list:
        jsonData['optional_args'].update({f'{i}': ''})
        jsonData['required_args'].pop(i)
    return jsonData

#======================================================
# Function to loop through site_groups for sensitve vars
#======================================================
def sensitive_var_site_group(**kwargs):
    if re.search('Grp_[A-F]', kwargs['site_group']):
        site_group = ast.literal_eval(os.environ[kwargs['site_group']])
        for x in range(1, 16):
            if not site_group[f'site_{x}'] == None:
                site_id = 'site_id_%s' % (site_group[f'site_{x}'])
                site_dict = ast.literal_eval(os.environ[site_id])
                if site_dict['run_location'] == 'local':
                    sensitive_var_value(**kwargs)
    else:
        site_id = 'site_id_%s' % (kwargs['site_group'])
        site_dict = ast.literal_eval(os.environ[site_id])
        if site_dict['run_location'] == 'local':
            sensitive_var_value(**kwargs)

#======================================================
# Function to add sensitive_var to Environment
#======================================================
def sensitive_var_value(**kwargs):
    sensitive_var = 'TF_VAR_%s' % (kwargs['Variable'])
    # -------------------------------------------------------------------------------------------------------------------------
    # Check to see if the Variable is already set in the Environment, and if not prompt the user for Input.
    #--------------------------------------------------------------------------------------------------------------------------
    if os.environ.get(sensitive_var) is None:
        print(f"\n----------------------------------------------------------------------------------\n")
        print(f"  The Script did not find {sensitive_var} as an 'environment' variable.")
        print(f"  To not be prompted for the value of {kwargs['Variable']} each time")
        print(f"  add the following to your local environemnt:\n")
        print(f"   - export {sensitive_var}='{kwargs['Variable']}_value'")
        print(f"\n----------------------------------------------------------------------------------\n")

    if os.environ.get(sensitive_var) is None:
        valid = False
        while valid == False:
            varValue = input('press enter to continue: ')
            if varValue == '':
                valid = True

        valid = False
        while valid == False:
            if kwargs.get('Multi_Line_Input'):
                print(f'Enter the value for {kwargs["Variable"]}:')
                lines = []
                while True:
                    # line = input('')
                    line = stdiomask.getpass(prompt='')
                    if line:
                        lines.append(line)
                    else:
                        break
                if not re.search('(certificate|private_key)', sensitive_var):
                    secure_value = '\\n'.join(lines)
                else:
                    secure_value = '\n'.join(lines)
            else:
                secure_value = stdiomask.getpass(prompt=f'Enter the value for {kwargs["Variable"]}: ')

            # Validate Sensitive Passwords
            cert_regex = re.compile(r'^\-{5}BEGIN (CERTIFICATE|PRIVATE KEY)\-{5}.*\-{5}END (CERTIFICATE|PRIVATE KEY)\-{5}$')
            if re.search('(certificate|private_key)', sensitive_var):
                if not re.search(cert_regex, secure_value):
                    valid = True
                else:
                    print(f'\n-------------------------------------------------------------------------------------------\n')
                    print(f'    Error!!! Invalid Value for the {sensitive_var}.  Please re-enter the {sensitive_var}.')
                    print(f'\n-------------------------------------------------------------------------------------------\n')
            elif re.search('(apikey|secretkey)', sensitive_var):
                if not sensitive_var == '':
                    valid = True
            else:
                if 'aes_passphrase' in sensitive_var:
                    sKey = 'aes_passphrase'
                    varTitle = 'Global AES Phassphrase'
                elif 'ntp_key' in sensitive_var:
                    sKey = 'key'
                    varTitle = 'NTP Key'
                elif re.search('snmp_(authorization|privacy)_key', sensitive_var):
                    sKey = 'snmp_key'
                    x = sensitive_var.split('_')
                    varType = '%s %s' % (x[0].capitalize(), x[1].capitalize())
                    varTitle = f'{varType}'
                elif 'radius_key' in sensitive_var:
                    sKey = 'radius_key'
                    varTitle = 'The RADIUS shared secret cannot contain backslashes, space, or hashtag "#".'
                elif 'radius_monitoring_password' in sensitive_var:
                    sKey = 'radius_monitoring_password'
                    varTitle = 'RADIUS Monitoring Password.'
                elif 'snmp_community' in sensitive_var:
                    sKey = 'snmp_community'
                    varTitle = 'The Community may only contain letters, numbers and the special characters of \
                    underscore (_), hyphen (-), or period (.). The Community may not contain the @ symbol.'
                elif 'smtp_password' in sensitive_var:
                    sKey = 'smtp_password'
                    varTitle = 'Smart CallHome SMTP Server Password'
                elif 'tacacs_key' in sensitive_var:
                    sKey = 'tacacs_key'
                    varTitle = 'The TACACS+ shared secret cannot contain backslashes, space, or hashtag "#".'
                elif 'tacacs_monitoring_password' in sensitive_var:
                    sKey = 'tacacs_monitoring_password'
                    varTitle = 'TACACS+ Monitoring Password.'
                elif 'vmm_password' in sensitive_var:
                    sKey = 'vmm_password'
                    varTitle = 'Virtual Networking Controller Password.'
                else:
                    print(sensitive_var)
                    print('Could not Match Sensitive Value Type')
                    exit()
                minimum = kwargs['jsonData'][sKey]['minimum']
                maximum = kwargs['jsonData'][sKey]['maximum']
                pattern = kwargs['jsonData'][sKey]['pattern']
                valid = validating.length_and_regex_sensitive(pattern, varTitle, secure_value, minimum, maximum)

        # Add the Variable to the Environment
        os.environ[sensitive_var] = '%s' % (secure_value)
        var_value = secure_value

    else:
        # Add the Variable to the Environment
        if kwargs.get('Multi_Line_Input'):
            var_value = os.environ.get(sensitive_var)
            var_value = var_value.replace('\n', '\\n')
        else:
            var_value = os.environ.get(sensitive_var)

    return var_value

#======================================================
# Function to Define stdout_log output
#======================================================
def stdout_log(ws, row_num, spot):
    if log_level == 0:
        return
    elif ((log_level == (1) or log_level == (2)) and
            (ws) and (row_num is None)) and spot == 'begin':
        print(f'-----------------------------------------------------------------------------\n')
        print(f'   Begin Worksheet "{ws.title}" evaluation...')
        print(f'\n-----------------------------------------------------------------------------\n')
    elif (log_level == (1) or log_level == (2)) and spot == 'end':
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   Completed Worksheet "{ws.title}" evaluation...')
        print(f'\n-----------------------------------------------------------------------------')
    elif log_level == (2) and (ws) and (row_num is not None):
        if re.fullmatch('[0-9]', str(row_num)):
            print(f'    - Evaluating Row   {row_num}...')
        elif re.fullmatch('[0-9][0-9]',  str(row_num)):
            print(f'    - Evaluating Row  {row_num}...')
        elif re.fullmatch('[0-9][0-9][0-9]',  str(row_num)):
            print(f'    - Evaluating Row {row_num}...')
    else:
        return

#======================================================
# Function to Validate Worksheet User Input
#======================================================
def validate_args(jsonData, **kwargs):
    globalData = kwargs['easy_jsonData']['components']['schemas']['globalData']['allOf'][1]['properties']
    global_args = [
        'admin_state',
        'application_epg',
        'application_profile',
        'annotation',
        'audit_logs',
        'cdp_interface_policy',
        'description',
        'events',
        'faults',
        'global_alias',
        'lldp_interface_policy',
        'management_epg',
        'management_epg_type',
        'monitoring_policy',
        'name',
        'name_alias',
        'pod_id',
        'port_channel_policy',
        'qos_class',
        'session_logs',
        'tenant',
        'username',
    ]
    for i in jsonData['required_args']:
        if i in global_args:
            if globalData[i]['type'] == 'integer':
                if kwargs[i] == None:
                    kwargs[i] = globalData[i]['default']
                else:
                    validating.number_check(i, globalData, **kwargs)
            elif globalData[i]['type'] == 'list_of_values':
                if kwargs[i] == None:
                    kwargs[i] = globalData[i]['default']
                else:
                    validating.list_values(i, globalData, **kwargs)
            elif globalData[i]['type'] == 'string':
                if not kwargs[i] == None:
                    validating.string_pattern(i, globalData, **kwargs)
            else:
                print(f'error validating.  Type not found {i}. 1')
                exit()
        elif i == 'site_group':
            validating.site_group('site_group', **kwargs)
        elif jsonData[i]['type'] == 'hostname':
            if not kwargs[i] == None:
                count = 1
                for hostname in kwargs[i].split(','):
                    kwargs[f'{i}_{count}'] = hostname
                    if ':' in hostname:
                        validating.ip_address(f'{i}_{count}', **kwargs)
                    elif re.search('[a-z]', hostname, re.IGNORECASE):
                        validating.dns_name(f'{i}_{count}', **kwargs)
                    else:
                        validating.ip_address(f'{i}_{count}', **kwargs)
                    kwargs.pop(f'{i}_{count}')
                    count += 1
        elif jsonData[i]['type'] == 'email':
            if not kwargs[i] == None:
                validating.email(i, **kwargs)
        elif jsonData[i]['type'] == 'integer':
            if kwargs[i] == None:
                kwargs[i] = jsonData[i]['default']
            else:
                validating.number_check(i, jsonData, **kwargs)
        elif jsonData[i]['type'] == 'list_of_domains':
            if not kwargs[i] == None:
                count = 1
                for domain in kwargs[i]:
                    kwargs[f'domain_{count}'] = domain
                    validating.domain(f'domain_{count}', **kwargs)
                    kwargs.pop(f'domain_{count}')
                    count += 1
        elif jsonData[i]['type'] == 'list_of_hosts':
            if not kwargs[i] == None:
                count = 1
                for hostname in kwargs[i].split(','):
                    kwargs[f'{i}_{count}'] = hostname
                    if ':' in hostname:
                        validating.ip_address(f'{i}_{count}', **kwargs)
                    elif re.search('[a-z]', hostname, re.IGNORECASE):
                        validating.dns_name(f'{i}_{count}', **kwargs)
                    else:
                        validating.ip_address(f'{i}_{count}', **kwargs)
                    kwargs.pop(f'{i}_{count}')
                    count += 1
        elif jsonData[i]['type'] == 'list_of_integer':
            if kwargs[i] == None:
                kwargs[i] = jsonData[i]['default']
            else:
                validating.number_list(i, jsonData, **kwargs)
        elif jsonData[i]['type'] == 'list_of_string':
            if not kwargs[i] == None:
                validating.string_list(i, jsonData, **kwargs)
        elif jsonData[i]['type'] == 'list_of_values':
            if kwargs[i] == None:
                kwargs[i] = jsonData[i]['default']
            else:
                validating.list_values(i, jsonData, **kwargs)
        elif jsonData[i]['type'] == 'list_of_vlans':
            if not kwargs[i] == None:
                validating.vlans(i, **kwargs)
        elif jsonData[i]['type'] == 'string':
            if not kwargs[i] == None:
                validating.string_pattern(i, jsonData, **kwargs)
        else:
            print(f'error validating.  Type not found {i}. 2')
            exit()
    for i in jsonData['optional_args']:
        if not kwargs[i] == None:
            if i in global_args:
                validating.validator(i, **kwargs)
            elif jsonData[i]['type'] == 'domain':
                validating.domain(i, **kwargs)
            elif jsonData[i]['type'] == 'email':
                validating.email(i, **kwargs)
            elif jsonData[i]['type'] == 'integer':
                validating.number_check(i, jsonData, **kwargs)
            elif jsonData[i]['type'] == 'list_of_values':
                validating.list_values(i, jsonData, **kwargs)
            elif jsonData[i]['type'] == 'list_of_vlans':
                if not kwargs[i] == None:
                    validating.vlans(i, **kwargs)
            elif jsonData[i]['type'] == 'list_of_string':
                if not kwargs[i] == None:
                    validating.string_list(i, jsonData, **kwargs)
            elif jsonData[i]['type'] == 'phone_number':
                validating.phone_number(i, **kwargs)
            elif jsonData[i]['type'] == 'string':
                validating.string_pattern(i, jsonData, **kwargs)
            else:
                print(f'error validating.  Type not found {i}. 3.')
                exit()

#======================================================
# Function to pull variables from easy_jsonData
#======================================================
def variablesFromAPI(**templateVars):
    valid = False
    while valid == False:
        json_vars = templateVars["jsonVars"]
        if 'popList' in templateVars:
            if len(templateVars["popList"]) > 0:
                for x in templateVars["popList"]:
                    varsCount = len(json_vars)
                    for r in range(0, varsCount):
                        if json_vars[r] == x:
                            json_vars.pop(r)
                            break
        print(f'\n-------------------------------------------------------------------------------------------\n')
        newDescr = templateVars["var_description"]
        if '\n' in newDescr:
            newDescr = newDescr.split('\n')
            for line in newDescr:
                if '*' in line:
                    print(fill(f'{line}',width=88, subsequent_indent='    '))
                else:
                    print(fill(f'{line}',88))
        else:
            print(fill(f'{templateVars["var_description"]}',88))
        print(f'\n    Select an Option Below:')
        for index, value in enumerate(json_vars):
            index += 1
            if value == templateVars["defaultVar"]:
                defaultIndex = index
            if index < 10:
                print(f'     {index}. {value}')
            else:
                print(f'    {index}. {value}')
        print(f'\n-------------------------------------------------------------------------------------------\n')
        if templateVars["multi_select"] == True:
            if not templateVars["defaultVar"] == '':
                var_selection = input(f'Please Enter the Option Number(s) to Select for {templateVars["varType"]}.  [{defaultIndex}]: ')
            else:
                var_selection = input(f'Please Enter the Option Number(s) to Select for {templateVars["varType"]}: ')
        else:
            if not templateVars["defaultVar"] == '':
                var_selection = input(f'Please Enter the Option Number to Select for {templateVars["varType"]}.  [{defaultIndex}]: ')
            else:
                var_selection = input(f'Please Enter the Option Number to Select for {templateVars["varType"]}: ')
        if not templateVars["defaultVar"] == '' and var_selection == '':
            var_selection = defaultIndex

        if templateVars["multi_select"] == False and re.search(r'^[0-9]+$', str(var_selection)):
            for index, value in enumerate(json_vars):
                index += 1
                if int(var_selection) == index:
                    selection = value
                    valid = True
        elif templateVars["multi_select"] == True and re.search(r'(^[0-9]+$|^[0-9\-,]+[0-9]$)', str(var_selection)):
            var_list = vlan_list_full(var_selection)
            var_length = int(len(var_list))
            var_count = 0
            selection = []
            for index, value in enumerate(json_vars):
                index += 1
                for vars in var_list:
                    if int(vars) == index:
                        var_count += 1
                        selection.append(value)
            if var_count == var_length:
                valid = True
            else:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  The list of Vars {var_list} did not match the available list.')
                print(f'\n-------------------------------------------------------------------------------------------\n')
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'  Error!! Invalid Selection.  Please Select a valid Option from the List.')
            print(f'\n-------------------------------------------------------------------------------------------\n')
    return selection

#======================================================
# Function to pull variables from easy_jsonData
#======================================================
def varBoolLoop(**templateVars):
    print(f'\n-------------------------------------------------------------------------------------------\n')
    newDescr = templateVars["Description"]
    if '\n' in newDescr:
        newDescr = newDescr.split('\n')
        for line in newDescr:
            if '*' in line:
                print(fill(f'{line}',width=88, subsequent_indent='    '))
            else:
                print(fill(f'{line}',88))
    else:
        print(fill(f'{templateVars["Description"]}',88))
    print(f'\n-------------------------------------------------------------------------------------------\n')
    valid = False
    while valid == False:
        varValue = input(f'{templateVars["varInput"]}  [{templateVars["varDefault"]}]: ')
        if varValue == '':
            if templateVars["varDefault"] == 'Y':
                varValue = True
            elif templateVars["varDefault"] == 'N':
                varValue = False
            valid = True
        elif varValue == 'N':
            varValue = False
            valid = True
        elif varValue == 'Y':
            varValue = True
            valid = True
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'   {templateVars["varName"]} value of "{varValue}" is Invalid!!! Please enter "Y" or "N".')
            print(f'\n-------------------------------------------------------------------------------------------\n')
    return varValue

#======================================================
# Function to pull variables from easy_jsonData
#======================================================
def varNumberLoop(**templateVars):
    maxNum = templateVars["maxNum"]
    minNum = templateVars["minNum"]
    varName = templateVars["varName"]

    print(f'\n-------------------------------------------------------------------------------------------\n')
    newDescr = templateVars["Description"]
    if '\n' in newDescr:
        newDescr = newDescr.split('\n')
        for line in newDescr:
            if '*' in line:
                print(fill(f'{line}',width=88, subsequent_indent='    '))
            else:
                print(fill(f'{line}',88))
    else:
        print(fill(f'{templateVars["Description"]}',88))
    print(f'\n-------------------------------------------------------------------------------------------\n')
    valid = False
    while valid == False:
        varValue = input(f'{templateVars["varInput"]}  [{templateVars["varDefault"]}]: ')
        if varValue == '':
            varValue = templateVars["varDefault"]
        if re.fullmatch(r'^[0-9]+$', str(varValue)):
            valid = validating.number_in_range(varName, varValue, minNum, maxNum)
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'   {varName} value of "{varValue}" is Invalid!!! ')
            print(f'   Valid range is {minNum} to {maxNum}.')
            print(f'\n-------------------------------------------------------------------------------------------\n')
    return varValue

#======================================================
# Function to pull variables from easy_jsonData
#======================================================
def varSensitiveStringLoop(**templateVars):
    maximum = templateVars["maximum"]
    minimum = templateVars["minimum"]
    varName = templateVars["varName"]
    varRegex = templateVars["varRegex"]

    print(f'\n-------------------------------------------------------------------------------------------\n')
    newDescr = templateVars["Description"]
    if '\n' in newDescr:
        newDescr = newDescr.split('\n')
        for line in newDescr:
            if '*' in line:
                print(fill(f'{line}',width=88, subsequent_indent='    '))
            else:
                print(fill(f'{line}',88))
    else:
        print(fill(f'{templateVars["Description"]}',88))
    print(f'\n-------------------------------------------------------------------------------------------\n')
    valid = False
    while valid == False:
        varValue = stdiomask.getpass(f'{templateVars["varInput"]} ')
        if not varValue == '':
            valid = validating.length_and_regex_sensitive(varRegex, varName, varValue, minimum, maximum)
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'   {varName} value is Invalid!!! ')
            print(f'\n-------------------------------------------------------------------------------------------\n')
    return varValue

#======================================================
# Function to pull variables from easy_jsonData
#======================================================
def varStringLoop(**templateVars):
    maximum = templateVars["maximum"]
    minimum = templateVars["minimum"]
    varName = templateVars["varName"]
    varRegex = templateVars["varRegex"]

    print(f'\n-------------------------------------------------------------------------------------------\n')
    newDescr = templateVars["Description"]
    if '\n' in newDescr:
        newDescr = newDescr.split('\n')
        for line in newDescr:
            if '*' in line:
                print(fill(f'{line}',width=88, subsequent_indent='    '))
            else:
                print(fill(f'{line}',88))
    else:
        print(fill(f'{templateVars["Description"]}',88))
    print(f'\n-------------------------------------------------------------------------------------------\n')
    valid = False
    while valid == False:
        varValue = input(f'{templateVars["varInput"]} ')
        if 'press enter to skip' in templateVars["varInput"] and varValue == '':
            valid = True
        elif not templateVars["varDefault"] == '' and varValue == '':
            varValue = templateVars["varDefault"]
            valid = True
        elif not varValue == '':
            valid = validating.length_and_regex(varRegex, varName, varValue, minimum, maximum)
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'   {varName} value of "{varValue}" is Invalid!!! ')
            print(f'\n-------------------------------------------------------------------------------------------\n')
    return varValue

#======================================================
# Function to Expand the VLAN list
#======================================================
def vlan_list_full(vlan_list):
    full_vlan_list = []
    if re.search(r',', str(vlan_list)):
        vlist = vlan_list.split(',')
        for v in vlist:
            if re.fullmatch('^\\d{1,4}\\-\\d{1,4}$', v):
                a,b = v.split('-')
                a = int(a)
                b = int(b)
                vrange = range(a,b+1)
                for vl in vrange:
                    full_vlan_list.append(int(vl))
            elif re.fullmatch('^\\d{1,4}$', v):
                full_vlan_list.append(int(v))
    elif re.search('\\-', str(vlan_list)):
        a,b = vlan_list.split('-')
        a = int(a)
        b = int(b)
        vrange = range(a,b+1)
        for v in vrange:
            full_vlan_list.append(int(v))
    else:
        full_vlan_list.append(vlan_list)
    return full_vlan_list

#======================================================
# Function to Expand a VLAN Range to a VLAN List
#======================================================
def vlan_range(vlan_list, **templateVars):
    results = 'unknown'
    while results == 'unknown':
        if re.search(',', str(vlan_list)):
            vx = vlan_list.split(',')
            for vrange in vx:
                if re.search('-', vrange):
                    vl = vrange.split('-')
                    min_ = int(vl[0])
                    max_ = int(vl[1])
                    if (int(templateVars['VLAN']) >= min_ and int(templateVars['VLAN']) <= max_):
                        results = 'true'
                        return results
                else:
                    if templateVars['VLAN'] == vrange:
                        results = 'true'
                        return results
            results = 'false'
            return results
        elif re.search('-', str(vlan_list)):
            vl = vlan_list.split('-')
            min_ = int(vl[0])
            max_ = int(vl[1])
            if (int(templateVars['VLAN']) >= min_ and int(templateVars['VLAN']) <= max_):
                results = 'true'
                return results
        else:
            if int(templateVars['VLAN']) == int(vlan_list):
                results = 'true'
                return results
        results = 'false'
        return results

#======================================================
# Function to Determine which sites to write files to.
#======================================================
def write_to_site(templateVars, **kwargs):
    class_type = templateVars['class_type']
    if re.search('(access|admin|fabric|site_policies|system_settings)', class_type):
        aci_template_path = pkg_resources.resource_filename(f'classes', 'templates/')
    else:
        aci_template_path = pkg_resources.resource_filename(f'class_{class_type}', 'templates/')

    templateLoader = jinja2.FileSystemLoader(
        searchpath=(aci_template_path + '%s/') % (class_type))
    templateEnv = jinja2.Environment(loader=templateLoader)
    ws = kwargs["ws"]
    row_num = kwargs["row_num"]
    site_group = str(kwargs['site_group'])
    
    # Define the Template Source
    kwargs["template"] = templateEnv.get_template(kwargs["template_file"])

    # Process the template
    if 'tenants' in class_type:
        kwargs["dest_dir"] = 'tenant_%s' % (templateVars['tenant'])
    else:
        kwargs["dest_dir"] = '%s' % (class_type)
    kwargs["dest_file"] = '%s.auto.tfvars' % (kwargs["tfvars_file"])
    if kwargs["initial_write"] == True:
        kwargs["write_method"] = 'w'
    else:
        kwargs["write_method"] = 'a'

    def process_siteDetails(site_dict, templateVars, **kwargs):
        # Create kwargs for site_name controller and controller_type
        kwargs['controller'] = site_dict.get('controller')
        kwargs['controller_type'] = site_dict.get('controller_type')
        templateVars['controller_type'] = site_dict.get('controller_type')
        kwargs['site_name'] = site_dict.get('site_name')
        kwargs['version'] = site_dict.get('version')
        if templateVars['template_type'] == 'firmware':
            templateVars['version'] = kwargs['version']
        if kwargs['controller_type'] == 'ndo' and templateVars['template_type'] == 'tenants':
            if templateVars['users'] == None:
                validating.error_tenant_users(**kwargs)
            else:
                for user in templateVars['users'].split(','):
                    regexp = '^[a-zA-Z0-9\_\-]+$'
                    validating.length_and_regex(regexp, 'users', user, 1, 64)
        # Create Terraform file from Template
        write_to_template(templateVars, **kwargs)

    if re.search('Grp_[A-F]', site_group):
        group_id = '%s' % (site_group)
        site_group = ast.literal_eval(os.environ[group_id])
        for x in range(1, 16):
            if not site_group[f'site_{x}'] == None:
                site_id = 'site_id_%s' % (site_group[f'site_{x}'])
                site_dict = ast.literal_eval(os.environ[site_id])

                # Add Site Detials to kwargs and write to template
                process_siteDetails(site_dict, templateVars, **kwargs)

    elif re.search(r'\d+', site_group):
        site_id = 'site_id_%s' % (site_group)
        site_dict = ast.literal_eval(os.environ[site_id])

        # Add Site Detials to kwargs and write to template
        process_siteDetails(site_dict, templateVars, **kwargs)

    else:
        print(f"\n-----------------------------------------------------------------------------\n")
        print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {kwargs['site_group']}.")
        print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
        print(f"\n-----------------------------------------------------------------------------\n")
        exit()

#======================================================
# Function to write files from Templates
#======================================================
def write_to_template(templateVars, **kwargs):
    opSystem  = platform.system()
    dest_dir  = kwargs["dest_dir"]
    dest_file = kwargs["dest_file"]
    site_name = kwargs["site_name"]
    template  = kwargs["template"]
    wr_method = kwargs["write_method"]

    if opSystem == 'Windows':
        if os.environ.get('TF_DEST_DIR') is None:
            tfDir = 'ACI'
        else:
            tfDir = os.environ.get('TF_DEST_DIR')
        if re.search(r'^\\.*\\$', tfDir):
            dest_dir = '%s%s\%s' % (tfDir, site_name, dest_dir)
        elif re.search(r'^\\.*\w', tfDir):
            dest_dir = '%s\%s\%s' % (tfDir, site_name, dest_dir)
        else:
            dest_dir = '.\%s\%s\%s' % (tfDir, site_name, dest_dir)
        if not os.path.isdir(dest_dir):
            mk_dir = 'mkdir %s' % (dest_dir)
            os.system(mk_dir)
        dest_file_path = '%s\%s' % (dest_dir, dest_file)
        if not os.path.isfile(dest_file_path):
            create_file = 'type nul >> %s' % (dest_file_path)
            os.system(create_file)
        tf_file = dest_file_path
        print(tf_file)
        wr_file = open(tf_file, wr_method)
    else:
        if os.environ.get('TF_DEST_DIR') is None:
            tfDir = 'ACI'
        else:
            tfDir = os.environ.get('TF_DEST_DIR')
        if re.search(r'^\/.*\/$', tfDir):
            dest_dir = '%s%s/%s' % (tfDir, site_name, dest_dir)
        elif re.search(r'^\/.*\w', tfDir):
            dest_dir = '%s/%s/%s' % (tfDir, site_name, dest_dir)
        else:
            dest_dir = './%s/%s/%s' % (tfDir, site_name, dest_dir)
        if not os.path.isdir(dest_dir):
            mk_dir = 'mkdir -p %s' % (dest_dir)
            os.system(mk_dir)
        dest_file_path = '%s/%s' % (dest_dir, dest_file)
        if not os.path.isfile(dest_file_path):
            create_file = 'touch %s' % (dest_file_path)
            os.system(create_file)
        tf_file = dest_file_path
        wr_file = open(tf_file, wr_method)

    # Remove Uneccessary Arguments

    # Render Payload and Write to File
    templateVars = json.loads(json.dumps(templateVars))
    if templateVars['class_type'] == 'system_settings':
        payload = template.render(templateVars)
    else:
        templateVars = {'keys':templateVars}
        payload = template.render(templateVars)
    wr_file.write(payload)
    wr_file.close()
